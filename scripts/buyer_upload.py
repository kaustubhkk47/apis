import openpyxl
import requests
import json

buyerURL = "http://api.probzip.com/users/buyer/"

redFill = openpyxl.styles.PatternFill(start_color='FA5858',end_color='FA5858',fill_type='solid')
greenFill = openpyxl.styles.PatternFill(start_color='00FF40',end_color='00FF40',fill_type='solid')

inputFileName = "./BuyerDataSheet.xlsx"
outputFileName = "./BuyerDataSheetF.xlsx"

def upload_buyers():
	wb = read_file()
	try:
		send_buyers_data(wb)
	except Exception as e:
		print e

def read_file():
	try:
		wb = openpyxl.load_workbook(inputFileName,data_only=True)
		print "File read correctly"	
	except Exception as e:
		print e
	return wb

def send_buyers_data(wb):
	row = 2
	column = 1
	ws = wb.worksheets[0]

	while True:
		buyerName = toString(ws.cell(row = row, column = column).value)
		if buyerName == "" or buyerName == None or buyerName == 'None':
			break
		buyerData = json.dumps(fill_buyer_data(wb, row))

		if not (buyerData == {} or buyerData == "{}"):
			response = requests.post(buyerURL, data = buyerData)
			if response.status_code == requests.codes.ok:
				try:
					jsonText = json.loads(response.text)

					if jsonText["statusCode"] != "2XX":
						post_feedback(wb, row, jsonText["body"]["error"])
						print jsonText["body"]["error"]
					else:
						post_feedback(wb, row, "Success",jsonText["body"]["buyer"]["buyerID"])
						print "Success"
				except Exception as e:
					print e
					post_feedback(wb, row, e)
			else:
				post_feedback(wb, row, "Error response from server")

		row += 1

	wb.save(outputFileName)

def fill_buyer_data(wb , i):
	buyerData = {}

	try:
		buyerData["name"] = toString(wb.worksheets[0]["A"+toString(i)].value)
		buyerData["company_name"] = toString(wb.worksheets[0]["B"+toString(i)].value)
		buyerData["mobile_number"] = toString(wb.worksheets[0]["C"+toString(i)].value)
		buyerData["email"] = toString(wb.worksheets[0]["D"+toString(i)].value)
		buyerData["password"] = toString(wb.worksheets[0]["E"+toString(i)].value)
		buyerData["alternate_phone_number"] = toString(wb.worksheets[0]["F"+toString(i)].value)
		buyerData["mobile_verification"] = True
		buyerData["email_verification"] = True
		buyerData["gender"] = toString(wb.worksheets[0]["G"+toString(i)].value)
	
		buyerDetails = {}
		buyerDetails["vat_tin"] = toString(wb.worksheets[0]["H"+toString(i)].value)
		buyerDetails["cst"] = toString(wb.worksheets[0]["I"+toString(i)].value)
		buyerDetails["buyer_interest"] = toString(wb.worksheets[0]["J"+toString(i)].value)
		buyerDetails["customer_type"] = toString(wb.worksheets[0]["K"+toString(i)].value)
		buyerDetails["purchase_duration"] = toString(wb.worksheets[0]["L"+toString(i)].value)
		buyerDetails["buying_capacity"] = toString(wb.worksheets[0]["M"+toString(i)].value)
		buyerDetails["buys_from"] = toString(wb.worksheets[0]["N"+toString(i)].value)
		buyerDetails["purchasing_states"] = toString(wb.worksheets[0]["O"+toString(i)].value)

		buyerData["details"] = buyerDetails

		buyerAddress = {}
		buyerAddress["address"] = toString(wb.worksheets[0]["P"+toString(i)].value)
		buyerAddress["landmark"] = toString(wb.worksheets[0]["Q"+toString(i)].value)
		buyerAddress["city"] = toString(wb.worksheets[0]["R"+toString(i)].value)
		buyerAddress["state"] = toString(wb.worksheets[0]["S"+toString(i)].value)
		buyerAddress["country"] = toString(wb.worksheets[0]["T"+toString(i)].value)
		buyerAddress["contact_number"] = toString(wb.worksheets[0]["U"+toString(i)].value)
		buyerAddress["pincode"] = toString(wb.worksheets[0]["V"+toString(i)].value)
		
		buyerData["address"] = buyerAddress

	except Exception as e:
		buyerData = {}
		print e
		print "Data was incorrect"
		post_feedback(wb, i, "Data was incorrect")

	return buyerData

def post_feedback(wb, i, feedback, buyerID=0):
	wb.worksheets[0]["X"+toString(i)].value = feedback
	if feedback == "Success":
		wb.worksheets[0]["X"+toString(i)].fill = greenFill
		wb.worksheets[0]["Y"+toString(i)] = int(buyerID)
	else:
		wb.worksheets[0]["X"+toString(i)].fill = redFill

def parseFloat(x):
	if x == None or x == 0:
		return float(0)
	else:
		return float(x)

def parseInt(x):
	if x == None or x == 0:
		return int(0)
	else:
		return int(x)

def toString(x):
	if (type(x) == "unicode"):
		x.encode("utf-8","ignore").strip()
	else:
		toString(x).strip()