## This script is useful only for downloading images where every a link has been given
## against every image
## The image links have to be kept in a workbook in sheet 1. Also create another workbook 2
## in which feedback is  given 

import urllib
import openpyxl
import os
import time

inputFileName = "./ImageDataSheet.xlsx"
outputFileName = "./ImageDataSheetf.xlsx"
foldername = "./images/"

def download_products():
	wb = read_file()
	try:
		download_product_images(wb)
	except Exception as e:
		print e

def read_file():
	wb = None
	try:
		print "Reading file"
		wb = openpyxl.load_workbook(inputFileName,data_only=True)
		print "File read correctly"	
	except Exception as e:
		print "File could not be read"
		print e
		raw_input("Press enter to exit")
		exit()
	return wb

def download_product_images(wb):
	row = 5
	column = 1
	ws = wb.worksheets[0]
	ws_feedback = wb.worksheets[1]
	create_image_directory()

	while True:
		imageLink = toString(ws.cell(row = row, column = column).value)
		if imageLink == "" or imageLink == None or imageLink == 'None':
			if(column==1):
				break
			else:
				row += 1
				column = 1
		else:
			imageName = str(row) + "-" + str(column) + ".jpg"
			urllib.urlretrieve(imageLink,foldername + imageName)
			ws_feedback.cell(row = row, column = column).value = imageName
			column += 1
			print "success"

	wb.save(outputFileName)

def toString(x):
	if (type(x) == "unicode"):
		return x.encode("utf-8","ignore").strip()
	else:
		return str(x).strip()

def create_image_directory():
	if not os.path.exists(foldername):
		os.makedirs(foldername)

if __name__ == "__main__":
	print "Enter 1 to download images"
	x = input("")
	if x == 1:
		download_products()
	else:
		print "Wrong input"
	raw_input("")