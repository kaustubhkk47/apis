import openpyxl
import requests
import json
import os
import shutil
import ast
from PIL import Image, ImageFile
import time

ImageFile.LOAD_TRUNCATED_IMAGES = True

productURL = "http://api.wholdus.com/products/"
redFill = openpyxl.styles.PatternFill(start_color='FA5858',end_color='FA5858',fill_type='solid')
greenFill = openpyxl.styles.PatternFill(start_color='00FF40',end_color='00FF40',fill_type='solid')

imageDirectory = "./images/"

allSizePaths = []
allImageSizes = [2000.0, 50.0, 100.0, 200.0, 300.0, 400.0, 600.00, 700.0]
for imageSize in allImageSizes:
	allSizePaths.append("{:.0f}x{:.0f}/".format(imageSize,imageSize))
allSizePaths[0] = ""
fileFormatExtensions = [".jpg", ".jpeg",".png"]

inputFileName = "./ProductDataSheet.xlsx"
outputFileName = "./ProductDataSheetf.xlsx"

startRow = 5

imageQualityPercent = 80

def upload_products():
	wb = read_file()
	success = 0
	try:
		success = send_products_data(wb)
	except Exception as e:
		print e
	return success

def modify_product_prices():
	wb = read_file()
	try:
		send_modified_product_prices(wb)
	except Exception as e:
		print e
	
def read_file():
	wb = None
	try:
		print "Reading file"
		wb = openpyxl.load_workbook(inputFileName,data_only=True)
		print "File read correctly"	
	except Exception as e:
		print "File could not be read"
		print e
		raw_input("Press enter to exit")
		exit()
	return wb

def send_products_data(wb):
	row = startRow
	column = 1
	ws = wb.worksheets[1]

	success = 1

	while True:
		productName = toString(ws.cell(row = row, column = column).value)
		if productName == "" or productName == None or productName == 'None':
			break
		productData = json.dumps(fill_product_data(wb, row))

		imageCount = countImages(row)

		if imageCount==0:
			post_feedback(wb, row, "0 images present")

		elif not (productData == {} or productData == "{}"):
			response = requests.post(productURL, data = productData)
			if response.status_code == requests.codes.ok:
				try:
					jsonText = json.loads(response.text)

					if jsonText["statusCode"] != "2XX":
						post_feedback(wb, row, jsonText["body"]["error"])
						print jsonText["body"]["error"]
						success = 0
					else:
						moveImages(jsonText, row, wb)
						post_feedback(wb, row, "Success",jsonText["body"]["product"]["productID"])
						print "Success"
				except Exception as e:
					print e
					post_feedback(wb, row, e)
					success = 0
			else:
				post_feedback(wb, row, "Error response from server")
				success = 0

		row += 1

	wb.save(outputFileName)
	return success

def send_modified_product_prices(wb):
	row = startRow
	column = 1
	ws = wb.worksheets[1]
	while True:
		productName = toString(ws.cell(row = row, column = column).value)
		if productName == "" or productName == None or productName == 'None':
			break
		productData = json.dumps(fill_modified_product_prices(wb, row))

		if not (productData == {} or productData == "{}"):
			response = requests.put(productURL, data = productData)
			if response.status_code == requests.codes.ok:
				try:
					jsonText = json.loads(response.text)

					if jsonText["statusCode"] != "2XX":
						post_feedback(wb, row, jsonText["body"]["error"])
						print jsonText["body"]["error"]
					else:
						post_feedback(wb, row, "Success",jsonText["body"]["product"]["productID"])
						print "Success"
				except Exception as e:
					print e
					post_feedback(wb, row, e)
			else:
				post_feedback(wb, row, "Error response from server")

		row += 1

	wb.save(outputFileName)

def moveImages(jsonText, row, wb):
	imgNo = 1
	body = jsonText["body"]["product"]["image"]
	image_path = toString(body["image_path"])
	image_name = toString(body["image_name"])
	image_numbers = toString(body["image_numbers"])
	if len(image_numbers) == 2:
		image_numbers = []
	elif len(image_numbers) == 3:
		image_numbers = [int(image_numbers[1])]
	else:
		image_numbers = toString(image_numbers[1:len(image_numbers)-1])
		image_numbers = [int(float(x)) for x in image_numbers.split(',')]
	for sizePath in allSizePaths:
		create_image_directory(image_path, sizePath)
	while True:
		check = 0
		for extension in fileFormatExtensions:
			imageFileName = toString(row) + "-" + toString(imgNo) + extension
			imagePath = os.path.join(imageDirectory, imageFileName)
			if os.path.isfile(imagePath):
				check = 1
				img = Image.open(imagePath)
				for i in range(len(allSizePaths)):
					sizePath = allSizePaths[i]
					directory = os.path.join(imageDirectory,image_path,sizePath)
					imageNewFileName = image_name + "-" + toString(image_numbers[imgNo-1]) + ".jpg"
					newPath = os.path.join(directory, imageNewFileName)
					imgnew = resize_image(img, allImageSizes[i])
					imgnew.save(newPath,format="JPEG",quality=imageQualityPercent,optimize=True, progressive=True)
				os.remove(imagePath)
				
		if(check == 0):
			break
		else:
			imgNo += 1
	post_image_feedback(wb, row, imgNo-1)

def create_image_directory(image_path, size_path):
	directory = os.path.join(imageDirectory,image_path,size_path)
	if not os.path.exists(directory):
		os.makedirs(directory)

def resize_image(img, x):
	width = img.width
	height = img.height
	if width > height and width > x:
		img = img.resize((int(x), int(x*float(height)/float(width))),Image.ANTIALIAS)
	if height >= width and height > x:
		img = img.resize((int(x*float(width)/float(height)), int(x)),Image.ANTIALIAS)
	return img
	
def post_image_feedback(wb, row, feedback):
	wb.worksheets[1]["V"+toString(row)].value = toString(feedback) + " images moved"

def fill_modified_product_prices(wb , i):
	productData = {}

	try:
		productData["productID"] = parseInt(wb.worksheets[1]["V"+toString(i)].value)
		productData["product_lot"] = fill_product_lot_data(wb,i)
	except Exception as e:
		post_feedback(wb, i, "Data was incorrect")

	return productData

def fill_product_data(wb , i):
	productData = {}

	try:
		productData["categoryID"] = get_categoryID(wb.worksheets[1]["D"+toString(i)].value, wb)
		productData["sellerID"] = parseInt(wb.worksheets[0]["F4"].value)

		productData["name"] = toString(wb.worksheets[1]["A"+toString(i)].value)
		productData["price_per_unit"] = parseFloat(wb.worksheets[2]["G"+toString(i)].value)
		productData["unit"] = toString(wb.worksheets[2]["E"+toString(i)].value)
		productData["tax"] = 0.0
		productData["lot_size"] = parseInt(wb.worksheets[2]["H"+toString(i)].value)
		productData["price_per_lot"] = parseFloat(productData["lot_size"]*productData["price_per_unit"])
		productData["image_count"] = countImages(i)
	
		productDetails = {}
		productDetails["seller_catalog_number"] = toString(wb.worksheets[1]["B"+toString(i)].value)
		productDetails["brand"] = toString(wb.worksheets[1]["C"+toString(i)].value)
		productDetails["description"] = toString(wb.worksheets[1]["E"+toString(i)].value)
		productDetails["gender"] = toString(wb.worksheets[1]["F"+toString(i)].value)
		productDetails["pattern"] = toString(wb.worksheets[1]["G"+toString(i)].value)
		productDetails["style"] = toString(wb.worksheets[1]["H"+toString(i)].value)
		productDetails["fabric_gsm"] = toString(wb.worksheets[1]["I"+toString(i)].value)
		productDetails["sleeve"] = toString(wb.worksheets[1]["J"+toString(i)].value)
		productDetails["neck_collar_type"] = toString(wb.worksheets[1]["K"+toString(i)].value)
		productDetails["length"] = toString(wb.worksheets[1]["L"+toString(i)].value)
		productDetails["work_decoration_type"] = toString(wb.worksheets[1]["M"+toString(i)].value)
		productDetails["colours"] = toString(wb.worksheets[1]["N"+toString(i)].value)
		productDetails["sizes"] = toString(wb.worksheets[1]["O"+toString(i)].value)
		productDetails["special_feature"] = toString(wb.worksheets[1]["P"+toString(i)].value)
		productDetails["packaging_details"] = toString(wb.worksheets[1]["Q"+toString(i)].value)
		productDetails["availability"] = toString(wb.worksheets[1]["R"+toString(i)].value)
		productDetails["weight_per_unit"] = parseFloat(wb.worksheets[2]["F"+toString(i)].value)
		productDetails["dispatched_in"] = toString(wb.worksheets[1]["S"+toString(i)].value)
		productDetails["sample_type"] = toString(wb.worksheets[2]["J"+toString(i)].value)
		productDetails["sample_description"] = toString(wb.worksheets[2]["K"+toString(i)].value)
		productDetails["sample_price"] = parseFloat(wb.worksheets[2]["L"+toString(i)].value)

		productDetails["manufactured_country"] = "India"
		productDetails["manufactured_city"] = toString(wb.worksheets[0]["C7"].value)

		productData["details"] = productDetails

		productData["product_lot"] = fill_product_lot_data(wb,i)

	except Exception as e:
		productData = {}
		print e
		print "Data was incorrect"
		post_feedback(wb, i, "Data was incorrect")

	return productData

def get_categoryID(value, wb):
	ws = wb.worksheets[3]
	row = 2
	maxRow = get_max_category_row(wb)
	for i in range(2,maxRow+1):
		if ws["A"+toString(i)].value == value:
			return parseInt(ws["B" +toString(i)].value)

def get_max_category_row(wb):
	row = 2
	ws = wb.worksheets[3]
	while True:
		categoryName = toString(ws["A"+toString(row)].value)
		if categoryName == "" or categoryName == None or categoryName == 'None':
			break
		else:
			row += 1
	return row-1

def countImages(row):
	imgNo = 1
	while True:
		check = 0
		for extension in fileFormatExtensions:
			imageFileName = toString(row) + "-" + toString(imgNo) + extension
			imagePath = os.path.join(imageDirectory,imageFileName)
			if os.path.isfile(imagePath):
				check = 1
		if(check == 0):
			break
		else:
			imgNo += 1
	return imgNo-1


def fill_product_lot_data(wb,i):
	productLotData = []

	column = 13
	ws = wb.worksheets[2]

	while True:
		from_lot = toString(ws.cell(row = i, column = column).value)
		if from_lot == 0 or from_lot == None or from_lot == "None":
			return productLotData
		to_lot = parseInt(ws.cell(row = i, column = column+1).value)
		price = parseFloat(ws.cell(row = i, column = column+2).value)

		productLot = {}
		productLot["lot_size_from"] = parseInt(from_lot)
		productLot["lot_size_to"] = to_lot
		productLot["price_per_unit"] = price
		productLotData.append(productLot)

		column += 3


	return productLotData

def post_feedback(wb, i, feedback, productID=0):
	wb.worksheets[1]["U"+toString(i)].value = feedback
	if feedback == "Success":
		wb.worksheets[1]["U"+toString(i)].fill = greenFill
		wb.worksheets[1]["V"+toString(i)] = int(productID)
	else:
		wb.worksheets[1]["U"+toString(i)].fill = redFill

def parseFloat(x):
	if x == None or x == 0:
		return float(0)
	else:
		return float(x)

def parseInt(x):
	if x == None or x == 0:
		return int(0)
	else:
		return int(x)

def toString(x):
	if (isinstance(x, unicode)):
		return x.encode("utf-8","ignore").strip()
	else:
		return str(x).strip()


if __name__ == "__main__":
	print "Enter 1 to upload products or 2 to modify product prices"
	x = input("")
	if x == 1:
		if(upload_products()==1):
			print "Products uploaded successfully"
		else:
			print "Some problem was encountered while uploading"
	elif x == 2:
		modify_product_prices()
	else:
		print "Wrong input"
	
	raw_input("Press enter to exit")