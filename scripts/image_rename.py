import openpyxl
import os
import string
import re
import shutil

inputFileName = "./ProductDataSheet.xlsx"
outputFileName = "./ProductDataSheetf.xlsx"

inputFolderName = "./OriginalImages/"
outputFolderName = "./images/"

redFill = openpyxl.styles.PatternFill(start_color='FA5858',end_color='FA5858',fill_type='solid')
greenFill = openpyxl.styles.PatternFill(start_color='00FF40',end_color='00FF40',fill_type='solid')	


def rename_images_inside_folders():
	checkDirectories()
	wb = read_file()
	try:
		start_rename_images_inside_folders(wb)
	except Exception as e:
		print e

def rename_images_with_filenames():
	checkDirectories()
	wb = read_file()
	try:
		start_rename_images_with_filenames(wb)
	except Exception as e:
		print e

def read_file():
	wb = None
	try:
		print "Reading file"
		wb = openpyxl.load_workbook(inputFileName,data_only=True)
		print "File read correctly"	
	except Exception as e:
		print "File could not be read"
		print e
		raw_input("Press enter to exit")
		exit()
	return wb

def checkDirectories():
	if os.path.exists(inputFolderName):
		pass
	else:
		print "Input folder not present"
		raw_input("Press enter to exit")
		exit()

def create_image_directory():
	if not os.path.exists(outputFolderName):
		os.makedirs(outputFolderName)

def start_rename_images_with_filenames(wb):
	row = 5
	column = 2
	ws = wb.worksheets[1]
	fileList = getAllFiles(inputFolderName)
	cleanFileList = getCleanedList(fileList)
	create_image_directory()

	while True:
		catalogNumber = toString(ws.cell(row = row, column = column).value)
		if catalogNumber == "" or catalogNumber == None or catalogNumber == 'None':
			break
		else:		
			count = 0

			for i in range(len(cleanFileList)):
				fileName = cleanFileList[i]
				if catalogNumber in fileName:
					filePath = os.path.join(inputFolderName,fileList[i])
					if (os.path.isfile(filePath)):
						imageNo = count + 1
						outputName = toString(row) + "-" + toString(imageNo) + ".jpg"
						newFilePath = os.path.join(outputFolderName,outputName)
						try:
							shutil.move(filePath,newFilePath)
							count += 1
						except:
							pass

			if count > 0:
				post_feedback(wb, row, "Success",count)
				print "Success"
			else:
				post_feedback(wb, row, "No images found",count)
				print "No images found"

		row += 1

	wb.save(outputFileName)

def start_rename_images_inside_folders(wb):
	row = 5
	column = 2
	ws = wb.worksheets[1]
	dirList = getAllDirectories()
	cleanDirList = getCleanedList(dirList)
	create_image_directory()

	while True:
		catalogNumber = toString(ws.cell(row = row, column = column).value)
		if catalogNumber == "" or catalogNumber == None or catalogNumber == 'None':
			break
		else:
			catalogNumber = removeSpecialChars(catalogNumber)
			catalogNumberPresent = 0
			try:
				index = cleanDirList.index(catalogNumber)
				catalogNumberPresent = 1
			except Exception as e:
				print e
				post_feedback(wb, row, "Catalog number not found")
			if (catalogNumberPresent == 1):
				imageDir = inputFolderName + dirList[index] + "/"
				moveImagesInsideFolders(imageDir,wb,row)
				
		row += 1

	wb.save(outputFileName)

def moveImagesInsideFolders(imageDir,wb,row):
	count = 0
	imageFiles = getAllFiles(imageDir)
	for f in imageFiles:
		try:
			imageNo = count + 1
			outputName = toString(row) + "-" + toString(imageNo) + ".jpg"
			fromFile = os.path.join(imageDir,f)
			toFile = os.path.join(outputFolderName,outputName)
			shutil.move(fromFile, toFile)
			count += 1
		except Exception as e:
			print e
			post_feedback(wb, row, "Problem while moving images",count)
			break

	if(count == len(imageFiles)):
		post_feedback(wb, row, "Success",count)
		shutil.rmtree(imageDir)
		print "Success"

def getAllFiles(imageDir):
	return [f for f in os.listdir(imageDir) if os.path.isfile(os.path.join(imageDir, f))]

def toString(x):
	if (isinstance(x, unicode)):
		return x.encode("utf-8","ignore").strip()
	else:
		return str(x).strip()

def removeSpecialChars(x):
	pattern = re.compile('[\W_]+')
	return pattern.sub("",x)

def getCleanedList(dirList):
	newList = []
	for dirName in dirList:
		newList.append(removeSpecialChars(dirName))
	return newList

def getAllDirectories():
	return next(os.walk(inputFolderName))[1]

def post_feedback(wb, i, feedback, count=0):
	wb.worksheets[1]["U"+toString(i)].value = feedback
	wb.worksheets[1]["V"+toString(i)] = toString(count) + "images renamed and moved"
	if feedback == "Success":
		wb.worksheets[1]["U"+toString(i)].fill = greenFill
	else:
		wb.worksheets[1]["U"+toString(i)].fill = redFill

if __name__ == "__main__":
	print "Enter 1 to rename images with folder names as catalog numbers and 2 to rename images with file names as catalog numbers"
	x = input("")
	if x == 1:
		rename_images_inside_folders()
	elif x == 2:
		rename_images_with_filenames()
	else:
		print "Wrong input"
	raw_input("Completed script. Press enter to exit")